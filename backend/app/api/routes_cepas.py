import asyncio
import csv
import io
import re
import unicodedata
from datetime import datetime, timezone
from typing import Annotated, Literal

import openpyxl
from litestar import Controller, get, post, patch, delete
from litestar.datastructures import UploadFile
from litestar.di import Provide
from litestar.enums import RequestEncodingType
from litestar.exceptions import NotFoundException, HTTPException
from litestar.params import Body
from litestar.status_codes import HTTP_204_NO_CONTENT
from pydantic import BaseModel

from app.schema.dtos import (
    CepaCreateDTO,
    CepaUpdateDTO,
    CepaResponseDTO,
    PaginatedCepasDTO,
    CepaFilterParams,
)
from app.repositories.cepa_repository import (
    CepaRepository,
    CepaNotFoundError,
    CepaAlreadyExistsError,
    resolve_coords_from_origen,
)
from app.models.models import Cepa
from app.core.security import admin_guard
from app.core.config import settings


def cepa_repository() -> CepaRepository:
    return CepaRepository()


class AddAttributeDTO(BaseModel):
    attribute_name: str
    values: dict[str, str | None]


# ---------------------------------------------------------------------------
# Import helpers
# ---------------------------------------------------------------------------

_MAX_FILE_MB = 10
_MAX_CEPA_ROWS = 5_000
_MAX_COLUMNS = 100
_MAX_VALUE_LEN = 1_000
_MAX_CEPA_LEN = 200

_XLSX_MAGIC = b"PK\x03\x04"
_XLS_MAGIC = b"\xd0\xcf\x11\xe0"
_RESERVED_KEYS = {"_id", "id", "embedding"}

# Campos fijos del modelo — el resto es dinámico
_FIXED_FIELD_ALIASES: dict[str, str] = {
    "cepa": "cepa",
    "latitud": "latitud",
    "longitud": "longitud",
    "envio_punta_arenas": "envio_punta_arenas",
    "envio punta arenas": "envio_punta_arenas",
    "envio a punta arenas": "envio_punta_arenas",
    "envío punta arenas": "envio_punta_arenas",
    "envío a punta arenas": "envio_punta_arenas",
}

_FLOAT_FIELDS = {"latitud", "longitud"}
_DATE_FIELDS = {"envio_punta_arenas"}
_NULL_VALUES = {"", "n/i", "n/a", "null", "none", "-"}


def _is_null(v: str) -> bool:
    return v.strip().lower() in _NULL_VALUES


def _check_magic(content: bytes, filename: str) -> None:
    magic = content[:4]
    if filename.endswith((".xlsx", ".xls")):
        if magic not in (_XLSX_MAGIC, _XLS_MAGIC):
            raise HTTPException(
                status_code=400,
                detail="El archivo no es un Excel válido. Asegurate de subir un archivo .xlsx o .xls real.",
            )
    else:
        if magic in (_XLSX_MAGIC, _XLS_MAGIC):
            raise HTTPException(
                status_code=400,
                detail="El archivo parece ser un Excel renombrado como .csv. Subilo con extensión .xlsx.",
            )


def _check_column_count(headers: list[str]) -> None:
    non_empty = [h for h in headers if h.strip()]
    if len(non_empty) > _MAX_COLUMNS:
        raise HTTPException(
            status_code=400,
            detail=f"El archivo tiene demasiadas columnas ({len(non_empty)}). El máximo permitido es {_MAX_COLUMNS}.",
        )


def _parse_date(v: str) -> datetime | None:
    v = v.strip()
    if not v or _is_null(v):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(v, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    return None


def _normalize_dynamic_key(header: str) -> str:
    """Convierte un encabezado a clave de campo MongoDB válida."""
    # Transliterate accented chars (ó→o, é→e, ñ→n, etc.) before stripping
    nfkd = unicodedata.normalize("NFKD", header.strip().lower())
    ascii_str = nfkd.encode("ascii", "ignore").decode("ascii")
    # Replace non-alphanum with '_', collapse runs, strip edges
    key = re.sub(r"[^a-zA-Z0-9_]", "_", ascii_str)
    key = re.sub(r"_+", "_", key)
    return key.strip("_")


# Pre-built lookup using normalized keys so headers with '/', '°', accents, etc. still match
_FIXED_FIELD_ALIASES_NORMALIZED: dict[str, str] = {
    _normalize_dynamic_key(k): v
    for k, v in _FIXED_FIELD_ALIASES.items()
}


def _parse_rows(raw_rows: list[dict[str, str]]) -> list[dict]:
    """Convierte las filas brutas del archivo a dicts listos para MongoDB."""
    parsed = []
    skipped_empty = 0
    for raw in raw_rows:
        doc: dict = {}
        for header, raw_value in raw.items():
            norm_key = _normalize_dynamic_key(header)
            field = _FIXED_FIELD_ALIASES_NORMALIZED.get(norm_key)

            if field == "cepa":
                val = raw_value.strip()[:_MAX_CEPA_LEN]
                if not val or _is_null(val):
                    continue
                doc["cepa"] = val

            elif field in _FLOAT_FIELDS:
                if _is_null(raw_value):
                    doc[field] = None
                else:
                    try:
                        doc[field] = float(raw_value.strip().replace(",", "."))
                    except ValueError:
                        doc[field] = None

            elif field in _DATE_FIELDS:
                doc[field] = _parse_date(raw_value)

            else:
                # Campo dinámico — reutiliza la clave ya normalizada
                if not norm_key:
                    continue
                if not re.match(r'^[a-zA-Z0-9_]+$', norm_key):
                    continue
                if norm_key in _RESERVED_KEYS:
                    continue
                val = None if _is_null(raw_value) else raw_value.strip()[:_MAX_VALUE_LEN]
                doc[norm_key] = val

        if doc.get("cepa"):
            lat, lon = resolve_coords_from_origen(
                doc.get("origen"), doc.get("latitud"), doc.get("longitud")
            )
            doc["latitud"] = lat
            doc["longitud"] = lon
            parsed.append(doc)
        else:
            skipped_empty += 1

    if settings.debug:
        print(f"[DEBUG][import] _parse_rows: {len(parsed)} válidas, {skipped_empty} sin cepa descartadas")
        if parsed:
            print(f"[DEBUG][import] ejemplo primer doc: {parsed[0]}")

    return parsed


def _read_csv(content: bytes) -> list[dict[str, str]]:
    # Intenta varios encodings — archivos de Excel suelen ser Latin-1/cp1252
    text: str | None = None
    used_encoding = "unknown"
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = content.decode(encoding)
            used_encoding = encoding
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("No se pudo decodificar el CSV (encodings probados: utf-8, latin-1, cp1252)")

    # Auto-detecta el delimitador (,  ;  \t  |)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ","

    raw_reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_csv_rows = list(raw_reader)

    if not all_csv_rows:
        return []

    headers = all_csv_rows[0]
    header_count = len(headers)
    _check_column_count(headers)

    rows: list[dict[str, str]] = []
    for line_num, row in enumerate(all_csv_rows[1:], start=2):
        if not any(v.strip() for v in row):
            continue
        if len(row) != header_count:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"El CSV tiene formato inválido en la fila {line_num}: "
                    f"se encontraron {len(row)} campo(s) pero el encabezado tiene {header_count}. "
                    f"Revisa que todos los delimitadores estén correctos."
                ),
            )
        rows.append(dict(zip(headers, row)))

    cepa_key = next((k for k in headers if k.strip().lower() == "cepa"), None)
    if cepa_key:
        cepa_count = sum(
            1 for r in rows
            if r.get(cepa_key, "").strip() and not _is_null(r[cepa_key])
        )
        if cepa_count > _MAX_CEPA_ROWS:
            raise HTTPException(
                status_code=400,
                detail=f"El archivo supera el límite de {_MAX_CEPA_ROWS:,} filas con cepa ({cepa_count:,} encontradas).",
            )

    if settings.debug:
        print(f"[DEBUG][import] CSV encoding={used_encoding!r}  delimiter={delimiter!r}  filas={len(rows)}  columnas={headers}")

    return rows


def _read_xlsx(content: bytes) -> list[dict[str, str]]:
    # Primera pasada: validar columnas y contar filas con cepa (streaming, barato)
    wb_pre = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if not wb_pre.worksheets:
        wb_pre.close()
        raise HTTPException(status_code=400, detail="El archivo Excel no contiene hojas.")
    ws_pre = wb_pre.worksheets[0]
    row_iter = ws_pre.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    if header_row is None:
        wb_pre.close()
        return []
    headers_lower = [str(h).strip().lower() if h is not None else "" for h in header_row]
    _check_column_count(headers_lower)
    cepa_idx = next((i for i, h in enumerate(headers_lower) if h == "cepa"), None)
    cepa_count = 0
    if cepa_idx is not None:
        for row in row_iter:
            if cepa_idx < len(row) and row[cepa_idx] is not None and str(row[cepa_idx]).strip():
                cepa_count += 1
    wb_pre.close()

    if cepa_count > _MAX_CEPA_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"El archivo supera el límite de {_MAX_CEPA_ROWS:,} filas con cepa ({cepa_count:,} encontradas).",
        )

    # Segunda pasada: carga completa para detectar filas ocultas
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    ws = wb.worksheets[0]
    all_rows = list(ws.iter_rows())
    if not all_rows:
        wb.close()
        return []
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in all_rows[0]]
    result = []
    for row in all_rows[1:]:
        if not row:
            continue
        excel_row_num = row[0].row  # número real de fila en Excel
        if ws.row_dimensions[excel_row_num].hidden:
            continue
        obj = {
            headers[i]: (str(cell.value) if cell.value is not None else "")
            for i, cell in enumerate(row)
            if i < len(headers) and headers[i]
        }
        if any(v.strip() for v in obj.values()):
            result.append(obj)
    wb.close()
    return result


class ImportRowResult(BaseModel):
    cepa: str
    status: Literal["created", "duplicate", "error"]
    error: str | None = None


class ImportResultDTO(BaseModel):
    created: int
    skipped: int
    errors: int
    rows: list[ImportRowResult]


async def _upsert_column_labels(new_labels: dict[str, str]) -> None:
    """Guarda/actualiza el mapa field_name→label original en MongoDB."""
    from app.models.models import ColumnLabels
    doc = await ColumnLabels.find_one(ColumnLabels.namespace == "cepas")
    if doc:
        doc.labels.update(new_labels)
        await doc.save()
    else:
        await ColumnLabels(namespace="cepas", labels=new_labels).insert()


def _collect_dynamic_labels(raw_rows: list[dict[str, str]]) -> dict[str, str]:
    """Devuelve {norm_key: header_original} para todos los campos dinámicos del archivo."""
    if not raw_rows:
        return {}
    labels: dict[str, str] = {}
    for original_header in raw_rows[0]:
        norm = _normalize_dynamic_key(original_header)
        if not norm or norm in _RESERVED_KEYS:
            continue
        if _FIXED_FIELD_ALIASES_NORMALIZED.get(norm):
            continue
        if re.match(r'^[a-zA-Z0-9_]+$', norm):
            labels[norm] = original_header.strip()
    return labels


# ---------------------------------------------------------------------------
# Controller
# ---------------------------------------------------------------------------

class CepaController(Controller):
    path = "/cepas"
    dependencies = {"repo": Provide(cepa_repository, sync_to_thread=False)}

    # ------------------------------------------------------------------
    # GET ALL
    # ------------------------------------------------------------------
    @get("/")
    async def get_all(
        self,
        repo: CepaRepository,
        cepa: str | None = None,
        offset: int = 0,
    ) -> PaginatedCepasDTO:
        filters = CepaFilterParams(cepa=cepa)
        items, total = await repo.get_all(filters)
        return PaginatedCepasDTO(
            total=total,
            items=[CepaResponseDTO(id=str(c.id), **c.model_dump(exclude={"id"})) for c in items],
        )

    # ------------------------------------------------------------------
    # GET /labels — mapa field→label para columnas dinámicas
    # ------------------------------------------------------------------
    @get("/labels")
    async def get_labels(self) -> dict[str, str]:
        from app.models.models import ColumnLabels
        doc = await ColumnLabels.find_one(ColumnLabels.namespace == "cepas")
        return doc.labels if doc else {}

    # ------------------------------------------------------------------
    # GET BY ID
    # ------------------------------------------------------------------
    @get("/{cepa_id:str}")
    async def get_by_id(self, cepa_id: str, repo: CepaRepository) -> CepaResponseDTO:
        try:
            cepa = await repo.get_by_id(cepa_id)
        except CepaNotFoundError as e:
            raise NotFoundException(detail=str(e))
        return CepaResponseDTO(id=str(cepa.id), **cepa.model_dump(exclude={"id"}))

    # ------------------------------------------------------------------
    # POST
    # ------------------------------------------------------------------
    @post("/", guards=[admin_guard])
    async def create(self, data: CepaCreateDTO, repo: CepaRepository) -> CepaResponseDTO:
        try:
            cepa = await repo.create(data)
        except CepaAlreadyExistsError as e:
            raise HTTPException(status_code=409, detail=str(e))
        return CepaResponseDTO(id=str(cepa.id), **cepa.model_dump(exclude={"id"}))

    # ------------------------------------------------------------------
    # POST /import — importa cepas desde CSV o Excel
    # ------------------------------------------------------------------
    @post("/import", guards=[admin_guard])
    async def import_from_file(
        self,
        data: Annotated[UploadFile, Body(media_type=RequestEncodingType.MULTI_PART)],
    ) -> ImportResultDTO:
        content = await data.read()
        filename = (data.filename or "").lower()

        if settings.debug:
            print(f"[DEBUG][import] archivo={data.filename!r}  tamaño={len(content)} bytes")

        if len(content) > _MAX_FILE_MB * 1024 * 1024:
            raise HTTPException(
                status_code=400,
                detail=f"El archivo supera el tamaño máximo de {_MAX_FILE_MB} MB.",
            )

        _check_magic(content, filename)

        if filename.endswith((".xlsx", ".xls")):
            raw_rows = _read_xlsx(content)
        else:
            raw_rows = _read_csv(content)

        if settings.debug:
            print(f"[DEBUG][import] filas brutas leídas: {len(raw_rows)}")

        if not raw_rows:
            raise HTTPException(status_code=400, detail="El archivo no contiene filas válidas")

        dynamic_labels = _collect_dynamic_labels(raw_rows)
        parsed_rows = _parse_rows(raw_rows)

        if settings.debug:
            print(f"[DEBUG][import] filas a insertar: {len(parsed_rows)}")

        if not parsed_rows:
            raise HTTPException(status_code=400, detail="No se encontraron filas con campo 'cepa' válido")

        seen_names: set[str] = set()
        dup_names: list[str] = []
        for doc in parsed_rows:
            norm = doc["cepa"].strip().lower()
            if norm in seen_names:
                if doc["cepa"] not in dup_names:
                    dup_names.append(doc["cepa"])
            else:
                seen_names.add(norm)
        if dup_names:
            sample = ", ".join(f'"{n}"' for n in dup_names[:10])
            suffix = f" (y {len(dup_names) - 10} más)" if len(dup_names) > 10 else ""
            raise HTTPException(
                status_code=400,
                detail=f"El archivo contiene cepas repetidas: {sample}{suffix}",
            )

        if dynamic_labels:
            await _upsert_column_labels(dynamic_labels)

        results: list[ImportRowResult] = []

        for doc in parsed_rows:
            cepa_name = doc.get("cepa", "")
            try:
                existing = await Cepa.find_one(Cepa.cepa == cepa_name)
                if existing:
                    if settings.debug:
                        print(f"[DEBUG][import] duplicada: {cepa_name!r}")
                    results.append(ImportRowResult(cepa=cepa_name, status="duplicate"))
                    continue

                cepa_obj = Cepa(**doc)
                await cepa_obj.insert()

                # Generar embedding si el módulo IA está disponible
                try:
                    from app.ia.services.chat.embedding_service import get_embedding_service
                    from app.ia.services.chat.dbSearch_service import DatabaseService
                    cepa_obj.embedding = await asyncio.to_thread(
                        get_embedding_service().encode,
                        DatabaseService._cepa_a_texto(cepa_obj),
                    )
                    await cepa_obj.save()
                except Exception:
                    pass

                if settings.debug:
                    print(f"[DEBUG][import] creada: {cepa_name!r}")
                results.append(ImportRowResult(cepa=cepa_name, status="created"))

            except Exception as e:
                if settings.debug:
                    print(f"[DEBUG][import] error en {cepa_name!r}: {e}")
                results.append(ImportRowResult(cepa=cepa_name, status="error", error=str(e)))

        return ImportResultDTO(
            created=sum(1 for r in results if r.status == "created"),
            skipped=sum(1 for r in results if r.status == "duplicate"),
            errors=sum(1 for r in results if r.status == "error"),
            rows=results,
        )

    # ------------------------------------------------------------------
    # POST /add-attribute
    # ------------------------------------------------------------------
    @post("/add-attribute", guards=[admin_guard])
    async def add_attribute(self, data: AddAttributeDTO) -> dict:
        field = data.attribute_name.strip()
        if not field:
            raise HTTPException(status_code=400, detail="attribute_name no puede estar vacío")
        if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', field):
            raise HTTPException(status_code=400, detail="attribute_name solo puede contener letras, números y guión bajo")

        updated = 0
        not_found = []

        for cepa_name, value in data.values.items():
            cepa = await Cepa.find_one(Cepa.cepa == cepa_name)
            if not cepa:
                not_found.append(cepa_name)
                continue
            await cepa.set({field: value, "fecha_actualizacion": datetime.now(timezone.utc)})
            updated += 1

        return {"updated": updated, "not_found": not_found}

    # ------------------------------------------------------------------
    # PATCH
    # ------------------------------------------------------------------
    @patch("/{cepa_id:str}", guards=[admin_guard])
    async def update(self, cepa_id: str, data: CepaUpdateDTO, repo: CepaRepository) -> CepaResponseDTO:
        try:
            cepa = await repo.update(cepa_id, data)
        except CepaNotFoundError as e:
            raise NotFoundException(detail=str(e))
        return CepaResponseDTO(id=str(cepa.id), **cepa.model_dump(exclude={"id"}))

    # ------------------------------------------------------------------
    # DELETE
    # ------------------------------------------------------------------
    @delete("/{cepa_id:str}", guards=[admin_guard], status_code=HTTP_204_NO_CONTENT)
    async def delete(self, cepa_id: str, repo: CepaRepository) -> None:
        try:
            await repo.delete(cepa_id)
        except CepaNotFoundError as e:
            raise NotFoundException(detail=str(e))
